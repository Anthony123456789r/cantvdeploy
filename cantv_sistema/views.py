from django.shortcuts import render, redirect, get_object_or_404
import matplotlib.pyplot as plt
from .models import *
import io
import traceback
import mpld3
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
from matplotlib.dates import DateFormatter
from django.contrib.auth import logout
from django.http import HttpResponse,HttpResponseForbidden
from django.core.files.base import ContentFile
from django.core.exceptions import ObjectDoesNotExist
from django.utils import timezone
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, landscape
import time
import re
from django.db.models import Q
from twilio.rest import Client
#==========LIBRERIAS DE AUTENTICACION DE IDENTIDAD=========
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from django.http import JsonResponse
# ========================================================
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.dimensions import ColumnDimension
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle,Paragraph,PageBreak
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import os
#===========
import datetime
import pytz
#===============
import qrcode
from io import BytesIO
#=============LIBRERIAS PARA EL BOT=============
import pyttsx3
import random
import threading
import cv2 
import numpy as np
from skimage.metrics import structural_similarity as ssim
from PIL import Image
#========================================================
# Create your views here.
#========FUNCION DE REPRODUCIR EL AUDIO DEL BOT========
def reproducir_audio():
    mensajes_bienvenida = [
        "¡Hola! Soy Lia, tu asistente virtual. Bienvenido al sistema.",
        "¡Hola! Soy Lia, tu asistente virtual. Bienvenida, por favor inicia sesión.",
        "¡Hola! Soy Lia, tu asistente virtual. Estamos felices de verte aquí.",
        "¡Hola! Soy Lia, tu asistente virtual. Estamos encantados de tenerte aquí en Canteve, la herramienta perfecta para crear contenido creativo.",
        "¡Hola! Soy Lia, tu asistente virtual. Bienvenido a Canteve, donde la magia de la creación se convierte en una poderosa herramienta en tus manos.",
        "¡Hola! Soy Lia, tu asistente virtual. En Canteve, la herramienta que hará brillar tu creatividad, te damos la bienvenida con los brazos abiertos."
    ]
    mensaje = random.choice(mensajes_bienvenida)
    iniciar1 = pyttsx3.init()
    voices = iniciar1.getProperty("voices")
    velocidad = 150
    iniciar1.setProperty("voice", voices[0].id)
    iniciar1.setProperty("rate", velocidad)
    iniciar1.say(mensaje)
    iniciar1.runAndWait()

def reproducir_audio_gestion_tecnico():
    mensajes_bienvenida = "Para dejar el reporte gestionado solo deberas llenar el formulario y estara listo"
    mensaje = random.choice(mensajes_bienvenida)
    iniciar1 = pyttsx3.init()
    voices = iniciar1.getProperty("voices")
    velocidad = 150
    iniciar1.setProperty("voice", voices[0].id)
    iniciar1.setProperty("rate", velocidad)
    iniciar1.say(mensaje)
    iniciar1.runAndWait()

def reproducir_audio_registro():
    mensajes_registro = [
        "¡Hola! Soy Lia, tu asistente virtual. Completa el formulario para unirte a nuestro equipo.",
        "¡Hola! Soy Lia, tu asistente virtual. Regístrate y forma parte de nuestro talentoso equipo de trabajo.",
        "¡Hola! Soy Lia, tu asistente virtual. ¡Únete a nosotros! Completa el formulario de registro para empezar.",
        "¡Hola! Soy Lia, tu asistente virtual. Queremos conocerte. Rellena el formulario y comienza una nueva etapa laboral.",
        "¡Hola! Soy Lia, tu asistente virtual. Inicia tu proceso de registro y descubre las oportunidades que tenemos para ti."
    ]
    mensaje = random.choice(mensajes_registro)
    iniciar3 = pyttsx3.init()
    voices = iniciar3.getProperty("voices")
    velocidad = 150
    iniciar3.setProperty("voice", voices[0].id)
    iniciar3.setProperty("rate", velocidad)
    iniciar3.say(mensaje)
    iniciar3.runAndWait()

def reproducir_audio_cliente():
    mensajes_bienvenida = [
        "¡Hola! Soy Lía, tu asistente virtual. ¡Bienvenido a nuestro sitio! Esperamos que disfrutes de tu experiencia aquí.",
        "¡Hola! Soy Lía, tu asistente virtual. Nos complace tenerte como nuestro cliente. Estamos aquí para ayudarte en lo que necesites.",
        "¡Hola! Soy Lía, tu asistente virtual. Bienvenido a nuestra plataforma. Estamos emocionados de ser parte de tu viaje.",
        "¡Hola! Soy Lía, tu asistente virtual. Gracias por elegirnos como tu proveedor de servicios. ¡Bienvenido a bordo!",
        "¡Hola! Soy Lía, tu asistente virtual. Te damos la bienvenida a nuestra comunidad. Esperamos brindarte un excelente servicio."
    ]
    mensaje = random.choice(mensajes_bienvenida)
    iniciar3 = pyttsx3.init()
    voices = iniciar3.getProperty("voices")
    velocidad = 150
    iniciar3.setProperty("voice", voices[0].id)
    iniciar3.setProperty("rate", velocidad)
    iniciar3.say(mensaje)
    iniciar3.runAndWait()
#========PAGINAS PRINCIPALES DEL SISTEMA========
def seleccion_login(request):
    if request.method == 'POST':
        seleccion = request.POST.get('login_select')
        if seleccion == 'cliente':
            return redirect("loginUser")
        elif seleccion == 'sistema':
            return redirect("loginSistema")
    audio_thread = threading.Thread(target=reproducir_audio)
    audio_thread.start()
    return render(request,"paginas_principales/login_seleccion.html")

#================================================

def registro(request):
     if request.method == 'POST':
        #======VARIABLES========
        rangoelegir = request.POST.get('rango')
        nombre = request.POST.get('usuario')
        apellido = request.POST.get('apellido')
        direccion = request.POST.get('direccion')
        cedula = request.POST.get('cedula')
        telefono = request.POST.get('telefono')
        #==========================================
        if nombre and apellido  and rangoelegir and direccion and cedula and telefono:
                if len(cedula) !=8:
                    return render(request,'paginas_principales/registro.html',{
                        "mensaje":"cedula incorrecta"
                        })
                elif len(telefono) !=11:
                    return render(request,'paginas_principales/registro.html',{
                        "mensaje":"telefono incorrecto"
                        })
                elif rangoelegir:
                    try:
                        validar_cod = rangos.objects.get(id=rangoelegir)
                        pass
                    except ObjectDoesNotExist:
                        return render(request,'paginas_principales/registro.html',{
                            "mensaje":"codigo invalido"
                            })
                    if rangoelegir == "1":
                        tecnicoNuevo = admin()
                        tecnicoNuevo.nombre = nombre
                        tecnicoNuevo.apellido =apellido
                        tecnicoNuevo.cedula =cedula
                        rango = rangos(id=1)
                        tecnicoNuevo.tipo_rango = rango
                        tecnicoNuevo.telefono = telefono
                        tecnicoNuevo.save()
                        return render(request,"paginas_principales/registro.html",{
                            "mensaje1": "Operacion exitosa y admin creado!"
                            })
                    elif rangoelegir == "2":
                        tecnicoNuevo = subAdmin()
                        tecnicoNuevo.nombre = nombre
                        tecnicoNuevo.apellido =apellido
                        tecnicoNuevo.cedula =cedula
                        rango = rangos(id=2)
                        tecnicoNuevo.tipo_rango = rango
                        tecnicoNuevo.telefono = telefono
                        tecnicoNuevo.save()
                        return render(request,"paginas_principales/registro.html",{
                            "mensaje1": "Operacion exitosa y subAdmin creado!"
                            })
                    elif rangoelegir == "3":
                        tecnicoNuevo = tecnico()
                        tecnicoNuevo.nombre = nombre
                        tecnicoNuevo.apellido =apellido
                        tecnicoNuevo.cedula =cedula
                        huella = ''
                        rango = rangos(id=3)
                        tecnicoEstado = estadoTecnico(id=1)
                        tecnicoNuevo.estado = tecnicoEstado
                        tecnicoNuevo.tipo_rango = rango
                        tecnicoNuevo.huella = huella
                        tecnicoNuevo.telefono = telefono
                        tecnicoNuevo.save()
                        return render(request,"paginas_principales/registro.html",{
                            "mensaje1": "Operacion exitosa y tecnico creado!"
                            })
        else:
            return render(request, 'paginas_principales/registro.html',{
                "mensaje": "todos los campos son obligatorios"
                })
     else:
        audio_thread = threading.Thread(target=reproducir_audio_registro)
        audio_thread.start()
        return render(request, 'paginas_principales/registro.html')
    #========VALIDACIONES DEL LOGIN==============
def identidad_user(request):
    if request.method == "POST":
        autenticar = request.POST.get('opcion')
        valTelefono = request.POST.get('valTelefono')
        valCorreo = request.POST.get('valCorreo')
        valGoogle = request.POST.get('valGoogle')
        valFacial = request.POST.get('valFacial')
        if autenticar == "telefono":
            # Credenciales de autenticación de Twilio
            account_sid = 'TWILIO_ACCOUNT_SID'
            auth_token = 'TWILIO_AUTH_TOKEN'

            # Generar un código de 4 dígitos
            codigo = str(random.randint(1000, 9999))

            # Número de teléfono de destino
            numero_destino = valTelefono  # Reemplaza con el número de teléfono correcto

            # Crear una instancia del cliente de Twilio
            client = Client(account_sid, auth_token)

            # Enviar el código de verificación al número de teléfono
            message = client.messages.create(
                body=f'Tu código de verificación es: {codigo}',
                from_='+16186346487',
                to=numero_destino
            )
            return render(request,"paginas_principales/validarUser.html")
        elif autenticar == "correo":
            try:
                validar_existencia = cliente.objects.get(correo=valCorreo)
                send_email()
            except cliente.DoesNotExist:
                   return render(request,"paginas_principales/validarUser.html",{
                       "mensaje":"Correo no existe"
                       })
            def send_email(sender_email, sender_password, recipient_email, subject, message):
                smtp_server = 'smtp.gmail.com'
                smtp_port = 587
                # Generar un código de 4 dígitos
                codigo = str(random.randint(1000, 9999))
                # Crear objeto MIMEMultipart
                msg = MIMEMultipart()
                msg['From'] = sender_email
                msg['To'] = recipient_email
                msg['Subject'] = subject

                # Agregar el cuerpo del mensaje
                msg.attach(MIMEText(message, 'plain'))

                try:
                    # Iniciar conexión SMTP
                    server = smtplib.SMTP(smtp_server, smtp_port)
                    server.starttls()

                    # Iniciar sesión en la cuenta de correo
                    server.login(sender_email, sender_password)

                    # Enviar el correo electrónico
                    server.send_message(msg)

                    # Cerrar conexión SMTP
                    server.quit()

                    print("Email sent successfully.")
                except Exception as e:
                    print("Failed to send email.")
                    print(str(e))

                    # Ejemplo de uso
                    sender_email = "Oneyver41@gmail.com"
                    sender_password = "kvbindjhvjtxbvel"
                    recipient_email = valCorreo
                    subject = "Correo de prueba"
                    message = f"Codigo de confirmacion {codigo}."

                    send_email(sender_email, sender_password, recipient_email, subject, message)
        elif autenticar == "google":
            pass
        elif autenticar == "facial":
            pass
        else:
            return render(request,"paginas_principales/validarUser.html",{
                "mensaje":"debe elegir una opcion de autenticacion valida"
                })         
    return render(request,"paginas_principales/validarUser.html")
def login_user(request):
    if request.method == 'POST':
        cedula = request.POST.get('cedula')
        probar_eliminar = cliente.objects.filter(eliminado=True,id=cedula)
        probar_eliminar1 = cliente.objects.filter(eliminado=True,cedula=cedula)
        if probar_eliminar or probar_eliminar1:
            return render(request, 'paginas_principales/login_user.html', {
                "mensaje": "Usuario inhabilitado"
                })
        if not cedula:
            return render(request, 'paginas_principales/login_user.html', {
                "mensaje": "Campo vacío"
                })
        try:
            dato = cliente.objects.get(cedula=cedula, tipo_rango=4)
            if dato:
                request.session["usuario"] = dato.nombre
                request.session["cedula"] = dato.cedula
                request.session["identificador"] = dato.id
                return redirect("cliente")
            else:
                return render(request, 'paginas_principales/login_user.html', {
                "mensaje": "Usuario no existe"
                })
        except cliente.DoesNotExist:
            pass
        try:
            dato = cliente.objects.get(id=cedula, tipo_rango=4)
            if dato:
                request.session["usuario"] = dato.nombre
                request.session["cedula"] = dato.cedula
                request.session["identificador"] = dato.id
                return redirect("cliente")
            else:
                return render(request, 'paginas_principales/login_user.html', {
                "mensaje": "Usuario no existe"
                })
        except cliente.DoesNotExist:
            return render(request, 'paginas_principales/login_user.html', {
                "mensaje": "Usuario no existe/código inválido"
                })
    else:
        audio_thread = threading.Thread(target=reproducir_audio_cliente)
        audio_thread.start()
        return render(request, 'paginas_principales/login_user.html')

    
def login(request):
    if request.method == 'POST':
        codigo_p00 = request.POST.get('p00')

        # Modelos a validar
        models_to_validate = [admin,subAdmin,tecnico]
        if not codigo_p00:
            return render(request, 'paginas_principales/login_user.html', {
                "mensaje": "Campo vacio"
                })
        
        for model in models_to_validate:
            try:
                objects = model.objects.get(id=codigo_p00,eliminado=False)
                if objects:
                        request.session["usuario"] = objects.nombre
                        request.session["cedula"] = objects.cedula
                        request.session["identificador"] = objects.id
                        if model == admin:
                            return redirect("admin")
                        elif model == subAdmin:
                            return redirect("subAdmin")
                        elif model == tecnico:
                            return redirect("tecnico")
            except model.DoesNotExist:
                pass

        return render(request, 'paginas_principales/login.html', {
            "mensaje": "usuario no existe/código inválido"
            })
    else:
        return render(request, 'paginas_principales/login.html')
    
def cerrarSesion_cantv(request):
     logout(request)
     return redirect("loginSistema")

def cerrarSesion_user(request):
     logout(request)
     return redirect("loginUser")
    #==========MODULO ADMIN=====
def modulo_admin(request):
    id_cli_sessi = request.session.get('identificador')
    try:
        compro_cli = admin.objects.get(id=id_cli_sessi)
        if compro_cli:
            pass
    except:
        return cerrarSesion_cantv(request)
    informacion = None
    mensaje = ""
    #===========VARIABLES DE INSTANCIAS=============
    clientesa = cliente.objects.filter(eliminado=False)
    empleados = tecnico.objects.filter(eliminado=False)
    subAdmins = subAdmin.objects.filter(eliminado=False)
    reportes_clientes = reporteCliente.objects.all()
    disp = estadoTecnico.objects.get(id=1)
    ocup = estadoTecnico.objects.get(id=2)
    emple_report = reportesParaTecnico.objects.filter(status=4).count() == 5
    emple_report_status_3 = reportesParaTecnico.objects.filter(status=3).count() == 5
    instan_status_report = estadoReporte.objects.get(id=2)
    #===================================================================
        
    if request.method == 'POST':
        estadisticas_fallas_report = request.POST.get('estad_report_select')
        busqueda = request.POST.get('busqueda')
        resultados = []
        informacion = []
        modelos = [admin, tecnico, cliente, subAdmin]
         # Iterar en los modelos
        
        if busqueda:
            for modelo in modelos:
                modelo_results = modelo.objects.filter(
                    Q(nombre__icontains=busqueda) |    # Comparar por nombre
                    Q(cedula__icontains=busqueda) |     # Comparar por cédula
                    Q(apellido__icontains=busqueda)|    # apellido
                    Q(id__icontains=busqueda)          # Comparar por ID
                )
                resultados.extend(modelo_results)
            if not resultados:
                mensaje = "sin resultados"

            # Realizar las acciones necesarias con los resultados
            for resultado in resultados:
                data = {
                    "id": resultado.id,
                    "nombre":resultado.nombre,
                    "apellido":resultado.apellido,
                    "cedula":resultado.cedula,
                    "rango_id":resultado.tipo_rango.id,
                    "rango":resultado.tipo_rango.rango,
                }
                informacion.append(data)
        else:
            mensaje = "campo vacio"
        if estadisticas_fallas_report:
            if estadisticas_fallas_report == "estad_fallas":
                reporte_cli_falla_proc = reporteCliente.objects.filter(status=instan_status_report).order_by('fecha')
                #========DATOS DE LA GRAFICA=========
                # Obtener las fechas de creación y contar la cantidad de reportes para cada fecha
                fechas = [reporte.fecha for reporte in reporte_cli_falla_proc]
                cantidades = list(range(1, len(reporte_cli_falla_proc) + 1))

                # Crear el gráfico de líneas
                plt.plot(fechas, cantidades)

                # Ajustar el formato del eje x para mostrar las fechas en línea
                plt.gca().xaxis.set_major_formatter(DateFormatter('%Y-%m-%d'))

                # Rotar las etiquetas del eje x para mejorar la legibilidad
                plt.xticks(rotation=10, ha='right')

                # Etiquetas y título
                plt.xlabel('Fecha de Creación')
                plt.ylabel('Cantidad de Reportes')
                plt.title('Cantidad de Reportes de Falla de Proceso a lo largo del tiempo')

                
               
                
    # ===================================================================================
    if not reportesParaTecnico.objects.exists() or emple_report or emple_report_status_3:
        tecnicos_asig = tecnico.objects.filter(estado=ocup if emple_report_status_3 else disp)
        tecnicos_asig.update(estado=disp)

        return render(request, 'modulos/admin/admin.html', {
            "clientes": clientesa, 
            "empleados": empleados, 
            "subAdmin": subAdmins,
            "search_results":informacion,
            "sinResult":mensaje,
            "report_cli_falla":reportes_clientes 
            })
    return render(request, 'modulos/admin/admin.html', {
        "clientes": clientesa, 
        "empleados": empleados, 
        "subAdmin": subAdmins,
        "search_results":informacion,
        "report_cli_falla":reportes_clientes          
        })
def obtener_datas_clientes(request):
        instan_cli = cliente.objects.all()
        instan_report = estadoReporte.objects.get(id=3)
        instan_data = reportesParaTecnico.objects.filter(status=instan_report)
        instan_data_peti = peticionesCliente.objects.all()
        return render(request, "modulos/admin/paginasAdmin/masDataCli.html",{
            "data_cli":instan_cli,
            "data_report_cli":instan_data,
            "data_peti":instan_data_peti
        })
def eliminacion_total_cliente(request,id):
    tiempo = datetime.datetime.now()
    instan_report = estadoReporte.objects.get(id=4)
    try:
        vali_user = cliente.objects.get(id=id)
        modo_ninja = RespaldoNinja(
            nombre = vali_user.nombre,
            apellido = vali_user.apellido,
            cedula = vali_user.cedula,
            fat = vali_user.fat1.fat,
            correo = vali_user.correo,
            cuadrilla = vali_user.cuadrilla,
            odn = vali_user.ond_fat.odn,
            telefono = vali_user.contacto,
            potencia_casa = vali_user.potencia_casa,
            potencia_fat = vali_user.potencia_fat,
            rango = vali_user.tipo_rango.rango,
            fecha_retiro = tiempo
        )
        modo_ninja.save()
        instancias = reportesParaTecnico.objects.filter(Datos_Cliente=id, status=instan_report)

        for instancia in instancias:
            # Realiza las operaciones que necesites con cada objeto
            instancia.DatosCliente_cedula = instancia.Datos_Cliente.cedula
            instancia.DatosCliente_nombre = instancia.Datos_Cliente.nombre
            instancia.DatosCliente_falla = instancia.DatosReporteCliente.categoria
            instancia.DatosCliente_FechaCreacion = instancia.DatosReporteCliente.fecha
            instancia.DatosCliente_id = instancia.Datos_Cliente.id
            instancia.save()

        vali_user.delete()
        return redirect("datos_cli_report")
    except cliente.DoesNotExist:
        pass
    return redirect("datos_cli_report")
def eliminar_permisos_admin(request):
    if request.method == "POST":
       id_capture = request.POST.get('id_capture')
       permiso_capture = request.POST.get('permiso_asignado')
       try:
        comprobar_data_elim = Permiso.objects.filter(
            Q(admin=id_capture)|
            Q(sub_admin=id_capture)|
            Q(tecnico=id_capture)|
            Q(cliente=id_capture),
            tipo_permiso=permiso_capture
        )
        comprobar_data_elim.delete()
        return redirect("permisosAdmin")
       except:
        pass
    return redirect("permisosAdmin")
def permisos_admin(request):
    id_de_permisos = None
    id_cli_sessi = request.session.get('identificador')
    try:
        compro_cli = admin.objects.get(id=id_cli_sessi)
        if compro_cli:
            pass
    except:
        return cerrarSesion_cantv(request)
    user_permiso = Permiso.objects.all()
    permisos_rango_por_usuario = {}
    
    for permiso in user_permiso:
        nombre_usuario = ""

        # Revisar qué campo de usuario está presente en el permiso
        if permiso.admin:
            nombre_usuario = permiso.admin.nombre
            rango_usuario = permiso.admin.tipo_rango.rango
            id_de_user = permiso.admin
        elif permiso.sub_admin:
            nombre_usuario = permiso.sub_admin.nombre
            rango_usuario = permiso.sub_admin.tipo_rango.rango
            id_de_user = permiso.sub_admin
        elif permiso.tecnico:
            nombre_usuario = permiso.tecnico.nombre
            rango_usuario = permiso.tecnico.tipo_rango.rango
            id_de_user = permiso.tecnico
        elif permiso.cliente:
            nombre_usuario = permiso.cliente.nombre
            rango_usuario = permiso.cliente.tipo_rango.rango
            id_de_user = permiso.cliente

        if nombre_usuario not in permisos_rango_por_usuario:
            permisos_rango_por_usuario[nombre_usuario] = {'permisos': [], 'rango': rango_usuario,"idUser":id_de_user}

        permisos_rango_por_usuario[nombre_usuario]['permisos'].append(permiso)
    todos_los_objetos = []
    for user in [subAdmin,tecnico,cliente]:
        objetos = user.objects.all()
        todos_los_objetos.extend(objetos)
    tipos_permisos = tiposPermisos.objects.all()
    #======================================================
    if request.method == 'POST':
        almacen_permisos = request.POST.getlist('permisos')
        id_de_permisos = request.POST.get('idUser')
        id_de_rango_user = int((request.POST.get('idRango')))

        if almacen_permisos:
            # Crear una instancia de Permiso para cada permiso seleccionado y guardarla en la base de datos
            if id_de_rango_user == 2:
                instan_sub = subAdmin.objects.get(id=id_de_permisos)
                for permiso_tipo in almacen_permisos:
                    tipo_permiso = tiposPermisos.objects.get(tipoPermiso=permiso_tipo)
                    permiso = Permiso.objects.create(sub_admin=instan_sub, tipo_permiso=tipo_permiso)
                    
            elif id_de_rango_user == 3:
                instan_tec = tecnico.objects.get(id=id_de_permisos)
                for permiso_tipo in almacen_permisos:
                    tipo_permiso = tiposPermisos.objects.get(tipoPermiso=permiso_tipo)
                    permiso = Permiso.objects.create(tecnico=instan_tec, tipo_permiso=tipo_permiso)
                    
            elif id_de_rango_user == 4:
                instan_cli = cliente.objects.get(id=id_de_permisos)
                for permiso_tipo in almacen_permisos:
                    tipo_permiso = tiposPermisos.objects.get(tipoPermiso=permiso_tipo)
                    permiso = Permiso.objects.create(cliente=instan_cli, tipo_permiso=tipo_permiso)
        return redirect("permisosAdmin")
    # Obtener los permisos ya asignados al usuario específico (subadmin, tecnico, cliente)
    permisos_usuario_actual = Permiso.objects.filter(
        Q(sub_admin__id=id_de_permisos) |
        Q(tecnico__id=id_de_permisos) |
        Q(cliente__id=id_de_permisos)
    )

    # Filtrar los permisos disponibles que aún no han sido asignados al usuario
    permisos_disponibles_usuario = tiposPermisos.objects.exclude(id__in=permisos_usuario_actual.values_list('tipo_permiso__id', flat=True))

    # Filtrar los permisos disponibles para el resto de los usuarios
    permisos_disponibles_admin = tiposPermisos.objects.exclude(id__in=Permiso.objects.filter(admin__isnull=False).values_list('tipo_permiso__id', flat=True))
    permisos_disponibles_subadmin = tiposPermisos.objects.exclude(id__in=Permiso.objects.filter(sub_admin__isnull=False).values_list('tipo_permiso__id', flat=True))
    permisos_disponibles_tecnico = tiposPermisos.objects.exclude(id__in=Permiso.objects.filter(tecnico__isnull=False).values_list('tipo_permiso__id', flat=True))
    permisos_disponibles_cliente = tiposPermisos.objects.exclude(id__in=Permiso.objects.filter(cliente__isnull=False).values_list('tipo_permiso__id', flat=True))

    return render(request, "modulos/admin/permisos.html",{
        "permisos":tipos_permisos,
        "usuarios":todos_los_objetos,
        "userPermi":user_permiso,
        'permisos_por_usuario': permisos_rango_por_usuario,
        'permisos_disponibles_usuario': permisos_disponibles_usuario,  # Pasar los permisos disponibles al contexto para el usuario actual
        'permisos_disponibles_admin': permisos_disponibles_admin,
        'permisos_disponibles_subadmin': permisos_disponibles_subadmin,
        'permisos_disponibles_tecnico': permisos_disponibles_tecnico,
        'permisos_disponibles_cliente': permisos_disponibles_cliente,
        })

#=========================================================
def registroODN(request):
    id_cli_sessi = request.session.get('identificador')
    try:
        compro_cli = admin.objects.get(id=id_cli_sessi)
        if compro_cli:
            pass
    except:
        return cerrarSesion_cantv(request)
    if request.method == "POST":
        odn = request.POST.get('odn')
        olt = request.POST.get('olt')
        direccion = request.POST.get('direccion')
        cantidadFat = request.POST.get('cantidadFAT')
        observacion = request.POST.get('observacionOdn')
        if odn and olt and direccion and cantidadFat:
            insertData = regisODN(
                odn = odn,
                olt = olt,
                cantidadFat = cantidadFat,
                direccion = direccion,
                observacion = observacion
            )
            insertData.save()
            return render(request, "modulos/admin/paginasAdmin/registrarODN.html",{
                "mensaje1":"Datos registrados"
                })
        else:
            return render(request, "modulos/admin/paginasAdmin/registrarODN.html",{
                "mensaje":"no pueden aver campos vacios"
                })
    return render(request, "modulos/admin/paginasAdmin/registrarODN.html")
# =========================================================
def obtener_datos_fat(request):
    id_cli_sessi = request.session.get('identificador')
    try:
        compro_cli = admin.objects.get(id=id_cli_sessi)
        if compro_cli:
            pass
    except:
        return cerrarSesion_cantv(request)
    if request.method == 'POST':
        odn_seleccionado = request.POST.get('odnSeleccionado')
        olt_options = []
        fat_options = []

        try:
            odn = regisODN.objects.get(odn=odn_seleccionado)
            olt_options = [odn.olt]
            fat_options = [f"-{str(i).zfill(2)}" for i in range(1, odn.cantidadFat + 1)]
        except regisODN.DoesNotExist:
            pass

        response = {
            'oltOptions': '<option selected disabled>Seleccione una opción</option>' + ''.join(
                f'<option value="{option}">{option}</option>' for option in olt_options),
            'fatOptions': '<option selected disabled>Seleccione una opción</option>' + ''.join(
                f'<option value="{option}">{option}</option>' for option in fat_options),
        }
        return JsonResponse(response)
    else:
        return JsonResponse({'oltOptions': '', 'fatOptions': ''})
# ===========================================================
def registroFAT(request):
    id_cli_sessi = request.session.get('identificador')
    try:
        compro_cli = admin.objects.get(id=id_cli_sessi)
        if compro_cli:
            pass
    except:
        return cerrarSesion_cantv(request)
    llamarodn = regisODN.objects.all()
    if request.method == "POST":
        direccionFAT = request.POST.get('direccionFat')
        fatElegir = request.POST.get('fatOpcion')
        observacionFAT = request.POST.get('observacionFat')
        odnElegir = request.POST.get('odnOpcion')
        if direccionFAT and fatElegir:
            instanODN = regisODN.objects.get(odn=odnElegir)
            registrarFATS = regisFAT(
                direccionFat = direccionFAT,
                observacionFat = observacionFAT,
                fat = fatElegir,
                dato_fat = instanODN
            )
            registrarFATS.save()
            return render(request, "modulos/admin/paginasAdmin/registrarFAT.html",{
                "llamado":llamarodn,
                "mensaje1":"Operacion exitosa"
                })
        else:
            return render(request, "modulos/admin/paginasAdmin/registrarFAT.html",{
                "llamado":llamarodn,"mensaje":"no puedne aver campos vacios"
                })
    return render(request, "modulos/admin/paginasAdmin/registrarFAT.html",{
        "llamado":llamarodn
        })

def editarCliente(request, id):
    id_cli_sessi = request.session.get('identificador')
    try:
        compro_cli = admin.objects.get(id=id_cli_sessi)
        if compro_cli:
            pass
    except:
        return cerrarSesion_cantv(request)
    cliente_obj = cliente.objects.get(id=id)
    cliente_existente = cliente.objects.filter(id=id)
    odn_llamado = regisODN.objects.all()
    if request.method == "POST":
        ond = request.POST.get('ond')
        nombre = request.POST.get('nombre')
        apellido = request.POST.get('apellido')
        identificador_servicio = request.POST.get('identificador_servicio')
        cedula = request.POST.get('cedula')
        contacto = request.POST.get('contacto')
        direccion = request.POST.get('direccion')
        cuadrilla = request.POST.get('cuadrilla')
        fat = request.POST.get('fat')
        posicion_fat = request.POST.get('posicion_fat')
        potencia_fat = request.POST.get('potencia_fat')
        potencia_casa = request.POST.get('potencia_casa')
        try:
            instan_odn_edit = regisODN.objects.get(odn=ond)
            instan_fat_edit = regisFAT.objects.get(fat=fat)
            if ond:
                cliente_obj.ond_fat = instan_odn_edit
            if fat:
                cliente_obj.fat1 = instan_fat_edit
        except:
            pass
        # Actualizar los campos solo si tienen información
        if nombre:
            cliente_obj.nombre = nombre
        if apellido:
            cliente_obj.apellido = apellido
        if identificador_servicio:
            cliente_obj.tipo_rango = identificador_servicio
        if cedula:
            cliente_obj.cedula = cedula
        if contacto:
            cliente_obj.contacto = contacto
        if direccion:
            cliente_obj.direccion = direccion
        if cuadrilla:
            cliente_obj.cuadrilla = cuadrilla
        if posicion_fat:
            cliente_obj.posicion_fat = posicion_fat
        if potencia_fat:
            cliente_obj.potencia_fat = potencia_fat
        if potencia_casa:
            cliente_obj.potencia_casa = potencia_casa

        cliente_obj.save()
        return render(request, 'modulos/admin/editar_cliente.html', {
            "cliente": cliente_obj,
              "cliente_existentes": cliente_existente,
                "mensaje1": "datos actualizados",
                "llamar_odn":odn_llamado
                })
    return render(request, 'modulos/admin/editar_cliente.html', {
        "cliente": cliente_obj,
          "cliente_existentes": cliente_existente,
          "llamar_odn":odn_llamado
          })
#============================
def mostrar_fats(request):
    if request.method == 'POST' and 'odnSeleccionado' in request.POST:
        odn_seleccionado = request.POST['odnSeleccionado']
        try:
            # Check if the selected ODN exists in the regisODN model
            odn_obj = regisODN.objects.get(odn=odn_seleccionado)
            # Filtrar los registros de regisFAT relacionados con el ODN seleccionado
            fats_relacionados = regisFAT.objects.filter(dato_fat=odn_obj)
            # Obtener los valores de los campos 'fat' de los registros relacionados
            fats = [f.fat for f in fats_relacionados]

            # Devolver la lista de FATS como respuesta JSON
            return JsonResponse({'fats': fats})

        except regisODN.DoesNotExist:
            # Si el ODN seleccionado no existe, devolver un JSON vacío
            return JsonResponse({'fats': []})

    # Si no se hizo una solicitud AJAX o no se encontró el parámetro 'odnSeleccionado', devolver un JSON vacío
    return JsonResponse({'fats': []})

def obtener_fats(request):
    odn_id = request.GET.get('odn_id')

    # Obtener los Fats según la ODN seleccionada
    fats = regisFAT.objects.filter(dato_fat_id=odn_id).values_list('fat', flat=True)

    return JsonResponse(list(fats), safe=False)
#===========================

def registroCliente(request):
    id_cli_sessi = request.session.get('identificador')
    try:
        compro_cli = admin.objects.get(id=id_cli_sessi)
        if compro_cli:
            pass
    except admin.DoesNotExist: # Es mejor especificar la excepción para evitar capturar errores inesperados
        return cerrarSesion_cantv(request) # Asegúrate de que esta función esté definida o importada

    llamarODN = regisODN.objects.all()
    llamarFAT = regisFAT.objects.all() # Aunque se cargan dinámicamente, sigue siendo útil pasarlas para la primera carga

    if request.method == 'POST':
        # ======VARIABLES========
        ond = request.POST.get('odn') # El nombre en el HTML es 'odn'
        usuario = request.POST.get('usuario')
        apellido = request.POST.get('apellido')
        ci = request.POST.get('ci')
        contacto = request.POST.get('contacto')
        direccion = request.POST.get('direccion')
        cuadrilla = request.POST.get('cuadrilla')
        fat = request.POST.get('fat')
        posicion_fat = request.POST.get('posicion_fat')
        potencia_fat = request.POST.get('potencia_fat')
        potencia_casa = request.POST.get('potencia_casa')
        servicio = request.POST.get('plan')
        correoCli = request.POST.get('correo')

        # ==========================================
        if all([ond, usuario, apellido, contacto, direccion, cuadrilla, fat,
                posicion_fat, potencia_fat, potencia_casa, ci, servicio, correoCli]):
            try:
                # Verificar si la cédula ya existe en la base de datos
                registrocliente_existente = cliente.objects.get(cedula=ci)
                return render(request, 'modulos/admin/registrar_cliente.html', {
                    "mensaje": "La cédula ya está registrada",
                    "llamadarODN": llamarODN,
                    "llamarFAT": llamarFAT,
                    "request_post": request.POST # Para mantener los datos del formulario si hay un error
                })
            except cliente.DoesNotExist:
                # Validar los datos de potencia (cambio de regex para aceptar cualquier número entero)
                if not re.match(r'^-?\d+$', potencia_fat) or not re.match(r'^-?\d+$', potencia_casa):
                    return render(request, 'modulos/admin/registrar_cliente.html', {
                        "mensaje": "Los datos de Potencia FAT y Potencia Casa deben ser números válidos.",
                        "llamadarODN": llamarODN,
                        "llamarFAT": llamarFAT,
                        "request_post": request.POST
                    })

                # Obtener la instancia del plan
                instanSer = None # Inicializa para evitar ReferenceError
                if servicio == "aba ultra 60mb":
                    instanSer = planes.objects.get(id=1)
                elif servicio == "aba ultra 100mb":
                    instanSer = planes.objects.get(id=2)
                elif servicio == "aba ultra 200mb":
                    instanSer = planes.objects.get(id=3)
                elif servicio == "aba ultra 300mb":
                    instanSer = planes.objects.get(id=4)
                else:
                    return render(request, 'modulos/admin/registrar_cliente.html', {
                        "mensaje": "Debe ingresar un tipo de servicio válido.",
                        "llamadarODN": llamarODN,
                        "llamarFAT": llamarFAT,
                        "request_post": request.POST
                    })

                try:
                    # Crear una instancia del modelo Registro
                    rango_cli = rangos.objects.get(id=4)
                    compro_odn = regisODN.objects.get(id=ond)
                    # CORRECCIÓN CLAVE: Usar 'id=fat' para buscar la instancia de regisFAT
                    compro_fat = regisFAT.objects.get(id=fat)

                    registrocliente = cliente(
                        ond_fat=compro_odn,
                        apellido=apellido,
                        nombre=usuario,
                        cedula=ci,
                        contacto=contacto,
                        direccion=direccion,
                        cuadrilla=cuadrilla,
                        fat1=compro_fat, # Asegúrate de que el campo en tu modelo 'cliente' se llama 'fat1'
                        posicion_fat=posicion_fat,
                        potencia_fat=potencia_fat,
                        potencia_casa=potencia_casa,
                        servicio=instanSer,
                        tipo_rango=rango_cli,
                        correo=correoCli
                    )
                    registrocliente.save()
                    return redirect('admin') # Asumo que 'admin' es el nombre de tu URL de éxito
                except Exception as e:
                    # Captura cualquier otro error durante el guardado (ej. ID de ODN/FAT no encontrado)
                    return render(request, 'modulos/admin/registrar_cliente.html', {
                        "mensaje": f"Ocurrió un error al guardar el cliente: {e}",
                        "llamadarODN": llamarODN,
                        "llamarFAT": llamarFAT,
                        "request_post": request.POST
                    })
        else:
            return render(request, 'modulos/admin/registrar_cliente.html', {
                "mensaje": "Todos los campos son obligatorios. Por favor, rellena el formulario completamente.",
                "llamadarODN": llamarODN,
                "llamarFAT": llamarFAT,
                "request_post": request.POST # Para mantener los datos del formulario
            })

    return render(request, 'modulos/admin/registrar_cliente.html', {
        "llamadarODN": llamarODN,
        "llamarFAT": llamarFAT,
        "request_post": request.POST # Pasa request.POST para que los campos se rellenen después de un refresh o error
    })

def hinabilitados(request):
    id_cli_sessi = request.session.get('identificador')
    try:
        compro_cli = admin.objects.get(id=id_cli_sessi)
        if compro_cli:
            pass
    except:
        return cerrarSesion_cantv(request)
    tecnico_hinabilitado = tecnico.objects.filter(eliminado=True)
    subAdmin_hinabilitado = subAdmin.objects.filter(eliminado=True)
    cliente_hinabilitado = cliente.objects.filter(eliminado=True)
    if request.method == 'POST':
        id_tecnico = request.POST.get('id_tecnico')
        id_subAdmin = request.POST.get('id_subAdmin')
        id_cliente = request.POST.get('id_cliente')
        if id_tecnico:
            info_tecnico = tecnico.objects.get(id=id_tecnico)
            info_tecnico.eliminado = False
            info_tecnico.save()

        elif id_subAdmin:
            info_subAdmin = subAdmin.objects.get(id=id_subAdmin)
            info_subAdmin.eliminado = False
            info_subAdmin.save()
        elif id_cliente:
            info_cliente = cliente.objects.get(id=id_cliente)
            info_cliente.eliminado = False
            info_cliente.save()

        return render(request, 'modulos/admin/inhabilitados.html',{
            "tecnicos":tecnico_hinabilitado,
            "sudAdmins":subAdmin_hinabilitado,
            "clientes":cliente_hinabilitado
            })
    else:
        return render(request, 'modulos/admin/inhabilitados.html',{
            "tecnicos":tecnico_hinabilitado,
            "sudAdmins":subAdmin_hinabilitado,
            "clientes":cliente_hinabilitado
            })
def historial_tecnico_pdf(request):
    tecnico_habilitado = tecnico.objects.exclude(eliminado=False)
    subAdmin_habilitado = subAdmin.objects.exclude(eliminado=False)

    if request.method == 'POST':
        id_tecnico = request.POST.get('id_tecnico_historial')
        report_tecnico = reportesParaTecnico.objects.filter(DatosTecnico=id_tecnico, status=4)

        # Crear una lista con los datos filtrados
        data = [['Tipo de Falla', 'Fecha Gestión', 'Información Técnico']]  # Encabezados de la tabla

        for reporte in report_tecnico:
            informacion_tecnico = reporte.mas_informacion_tecnico if reporte.mas_informacion_tecnico else 'SIN INFORMACIÓN'
            
            # Formatear la fecha de gestión a formato de 12 horas con "am" y "pm"
            fecha_gestion = reporte.fecha_gestion.strftime('%d/%m/%Y %I:%M %p')
            
            # Agregar los datos de cada reporte a la lista
            data.append([reporte.DatosReporteCliente.categoria, fecha_gestion, informacion_tecnico])

        # Crear un objeto response
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="historial_tecnico.pdf"'

        # Crear el documento PDF
        doc = SimpleDocTemplate(response, pagesize=letter)
        elements = []

        # Estilo para los párrafos
        styles = getSampleStyleSheet()
        normal_style = styles['Normal']

        # Crear la tabla con los datos
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  # Color de fondo para los encabezados
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),  # Color de texto para los encabezados
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Alinear contenido al centro
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Fuente en negrita para los encabezados
            ('FONTSIZE', (0, 0), (-1, 0), 12),  # Tamaño de fuente para los encabezados
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),  # Espacio inferior para los encabezados
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),  # Color de fondo para las celdas de datos
            ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Líneas de rejilla
            ('FONTSIZE', (0, 1), (-1, -1), 10),  # Tamaño de fuente para las celdas de datos
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),  # Espacio inferior para las celdas de datos
        ]))

        elements.append(Paragraph('Historial Técnico', normal_style))
        elements.append(table)

        # Agregar una nueva página si el contenido llena la página actual
        if table.wrap(doc.width, doc.height)[1] < doc.height:
            elements.append(PageBreak())

        doc.build(elements)

        return response

    return render(request, 'modulos/admin/inhabilitados.html', {
        "tecnicos": tecnico_habilitado, 
        "subAdmins": subAdmin_habilitado
        })
 
def ver_reportes_gestion(request):
    id_cli_sessi = request.session.get('identificador')
    try:
        compro_cli = admin.objects.get(id=id_cli_sessi)
        if compro_cli:
            pass
    except:
        return cerrarSesion_cantv(request)
    
    reporteGestion = reportesParaTecnico.objects.filter(status=4,Datos_Cliente__isnull=False,DatosReporteCliente__isnull=False)
    reporteGestion1 = reportesParaTecnico.objects.filter(status=4,Datos_Cliente__isnull=True,DatosReporteCliente__isnull=True)
    # Combinar los resultados de ambas consultas
    reportes_combinados = list(reporteGestion) + list(reporteGestion1)

    if 'export_excel' in request.GET:
        # Crear un nuevo libro de Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Agregar encabezados de columna
        headers = ["ID", "Nombre Cliente", "Tipo de Falla", "Cédula Cliente", "Fecha Reporte", "Nombre Técnico", "Status", "Informacion de tecnico", "Fecha de gestion"]
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            sheet[f"{column_letter}1"] = header
            sheet[f"{column_letter}1"].font = Font(bold=True)

        # Agregar datos de los reportes gestionados
        for row_num, reporte in enumerate(reportes_combinados, 2):
            sheet[f"A{row_num}"] = reporte.id
            if reporte.Datos_Cliente:
                sheet[f"B{row_num}"] = reporte.Datos_Cliente.nombre
                sheet[f"D{row_num}"] = reporte.Datos_Cliente.cedula
                fecha_reporte = reporte.DatosReporteCliente.fecha if reporte.DatosReporteCliente.fecha else timezone.now()
                fecha_reporte_str = fecha_reporte.strftime("%Y-%m-%d %H:%M:%S")
                sheet[f"E{row_num}"] = fecha_reporte_str
            else:
                sheet[f"B{row_num}"] = reporte.DatosCliente_nombre
                sheet[f"D{row_num}"] = reporte.DatosCliente_cedula
                fecha_reporte = reporte.DatosCliente_FechaCreacion if reporte.DatosCliente_FechaCreacion else timezone.now()
                fecha_reporte_str = fecha_reporte.strftime("%Y-%m-%d %H:%M:%S")
                sheet[f"E{row_num}"] = fecha_reporte_str

            sheet[f"C{row_num}"] = reporte.DatosReporteCliente.categoria if reporte.DatosReporteCliente else reporte.DatosCliente_falla
            sheet[f"F{row_num}"] = reporte.DatosTecnico.nombre
            sheet[f"G{row_num}"] = reporte.status.estadoReporter
            sheet[f"H{row_num}"] = reporte.mas_informacion_tecnico if reporte.mas_informacion_tecnico else ""
            sheet[f"I{row_num}"] = reporte.fecha_gestion if reporte.fecha_gestion else timezone.now()
            fecha_gestion = reporte.fecha_gestion if reporte.fecha_gestion else timezone.now()
            fecha_gestion_str = fecha_gestion.strftime("%Y-%m-%d %H:%M:%S")  # Formato de cadena sin zona horaria
            sheet[f"I{row_num}"] = fecha_gestion_str

        # Ajustar el ancho de las columnas al contenido
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) if max_length > len(header) else len(header)
            column_letter = column[0].column_letter
            sheet.column_dimensions[column_letter].width = adjusted_width

        # Guardar el libro de Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=reportes_gestionados.xlsx'
        workbook.save(response)
        return response

    return render(request, 'modulos/admin/paginasAdmin/reportesGestion.html', {
        "reportesGestionados": reportes_combinados
    })
    
def ver_reportes_tecnico(request, id):
    id_cli_sessi = request.session.get('identificador')
    try:
        compro_cli = admin.objects.get(id=id_cli_sessi)
        if compro_cli:
            pass
    except:
        return cerrarSesion_cantv(request)
    tecnico_identidad = tecnico.objects.get(id=id)
    paraTecnicoo = reportesParaTecnico.objects.filter(DatosTecnico=tecnico_identidad, status=3)
    
    if request.method == "POST":
        reporte_reasignar_id = request.POST.get('reporte_id')
        reporte_reasignar = reportesParaTecnico.objects.get(id=reporte_reasignar_id)
        tecnico_disponible = tecnico.objects.filter(estado__estado='disponible').first()

        if tecnico_disponible:
            # Obtener el técnico anterior y actualizar su estado si no tiene más informes asignados
            informes_asignados_anterior = reportesParaTecnico.objects.filter(DatosTecnico=tecnico_identidad).count()
            if informes_asignados_anterior == 1:
                tecnico_identidad.estado = estadoTecnico.objects.get(id=1)
                tecnico_identidad.save()

            # Actualizar el informe para incluir el nuevo técnico
            reporte_reasignar.DatosTecnico = tecnico_disponible
            reporte_reasignar.save()

            # Actualizar el estado del técnico nuevo si tiene cinco informes asignados
            informes_asignados = reportesParaTecnico.objects.filter(DatosTecnico=tecnico_disponible).count()
            if informes_asignados == 5:
                tecnico_disponible.estado = estadoTecnico.objects.get(id=2)
                tecnico_disponible.save()
                return redirect("admin")


    return render(request, 'modulos/admin/ver_reportes_tecnico.html', {
        "tecnico": tecnico_identidad,
        "nuevo_reporte": paraTecnicoo
          })

def enviar_reporte_tecnico(request, id):
    id_cli_sessi = request.session.get('identificador')
    try:
        compro_cli = admin.objects.get(id=id_cli_sessi)
        if compro_cli:
            pass
    except:
        return cerrarSesion_cantv(request)
    reporte_cliente = reporteCliente.objects.filter(valor=True)
    tecnico_enviar = tecnico.objects.get(id=id)
    paraTecnicoo = reportesParaTecnico.objects.filter(DatosTecnico=tecnico_enviar, status=3)

    if request.method == "POST":
        reporte_id = request.POST.get('reporte_id')
        tecnico_id = request.POST.get('tecnico_id')

        # ======instancias=============
        reporte_seleccionado = reporteCliente.objects.get(id=reporte_id)
        tecnico_seleccionado = tecnico.objects.get(id=tecnico_id)
        estado_reporte = estadoReporte.objects.get(estadoReporter="asignado")
        admin_sesion = request.session.get('cedula')
        admin_id = get_object_or_404(admin, cedula=admin_sesion)
        admin_id = admin_id.id
        admin_instancia = admin.objects.get(id=admin_id)
        # =================================

        informes_asignados = reportesParaTecnico.objects.filter(DatosTecnico=tecnico_seleccionado,status=estado_reporte).count()

        if tecnico_seleccionado.estado.estado == 'disponible' or informes_asignados < 5:
            #=====instancias======
            el_tipo = tipoReporte.objects.get(id=1)
            #=======================
            nuevo_reporte = reportesParaTecnico(
                DatosCliente=reporte_seleccionado,
                DatosReporteCliente=reporte_seleccionado,
                status=estado_reporte,
                DatosTecnico=tecnico_seleccionado,
                DatosAdmin=admin_instancia,
                tipo_reporte=el_tipo
            )
            informes_asignados += 1
            reporte_seleccionado.valor = False
            reporte_seleccionado.save()
            nuevo_reporte.save()
            if informes_asignados == 5:
                 tecnico_seleccionado.estado.estado = "ocupado"
                 tecnico_seleccionado.save()
                 return redirect('admin') 
            else:
                return render(request, 'modulos/admin/enviar_reporte_tecnico.html', {
                    "reporte": reporte_cliente, 
                    "tecnico": tecnico_enviar, 
                    "nuevo_reportee": paraTecnicoo
                    })
    else:
        return render(request, 'modulos/admin/enviar_reporte_tecnico.html', {
            "reporte": reporte_cliente, 
            "tecnico": tecnico_enviar, 
            "nuevo_reportee": paraTecnicoo
            })

def editarTecnico(request, id):
    id_cli_sessi = request.session.get('identificador')
    try:
        compro_cli = admin.objects.get(id=id_cli_sessi)
        if compro_cli:
            pass
    except:
        return cerrarSesion_cantv(request)
    tecnico_obj = tecnico.objects.get(id=id)
    empleado = tecnico.objects.filter(id=id)
    if request.method == "POST":
        nombre = request.POST.get("nombre")
        apellido = request.POST.get("apellido")
        direccion = request.POST.get("direccion")
        telefono = request.POST.get("telefono")
        
        if nombre:
            tecnico_obj.nombre = nombre
        if apellido:
            tecnico_obj.apellido = apellido
        if direccion:
            tecnico_obj.direccion = direccion
        if telefono:
            tecnico_obj.telefono = telefono
        tecnico_obj.save()
        
        return render(request, 'modulos/admin/editarTecnico.html', {
            "tecnico": tecnico_obj, 
            "empleados": empleado, 
            "mensaje": "Datos actualizados"
            })

    return render(request, 'modulos/admin/editarTecnico.html', {
        "tecnico": tecnico_obj, 
        "empleados": empleado
        })


def editarSubAdmin(request,id):
    id_cli_sessi = request.session.get('identificador')
    try:
        compro_cli = admin.objects.get(id=id_cli_sessi)
        if compro_cli:
            pass
    except:
        return cerrarSesion_cantv(request)
    subadmin = subAdmin.objects.get(id=id)
    empleados = subAdmin.objects.exclude(id=id)
    if request.method == "POST":
        nombre = request.POST.get("nombre")
        apellido = request.POST.get("apellido")
        direccion = request.POST.get("direccion")
        telefono = request.POST.get("telefono")
        
        if nombre:
            subadmin.nombre = nombre
        if apellido:
            subadmin.apellido = apellido
        if direccion:
            subadmin.direccion = direccion
        if telefono:
            subadmin.telefono = telefono
        subadmin.save()
        return render(request, 'modulos/admin/editarSubAdmin.html', {
            "subadmin": subadmin, 
            "empleados": empleados, 
            "mensaje1": "operacion exitosa"
            })
    else:
        return render(request, 'modulos/admin/editarSubAdmin.html', {
            "subadmin": subadmin, 
            "empleados": empleados
            })

def crearTecnico(request):
    id_cli_sessi = request.session.get('identificador')
    try:
        compro_cli = admin.objects.get(id=id_cli_sessi)
        if compro_cli:
            pass
    except:
        return cerrarSesion_cantv(request)
    if request.method == "POST":
        nombre = request.POST.get("empleado")
        apellido = request.POST.get("apellidoEmpleado")
        cedula = request.POST.get("cedulaEmpleado")
        direccion = request.POST.get("direccion")
        telefono = request.POST.get("telefono")
        if nombre and apellido and direccion and telefono and cedula:
                tecnicoNuevo = tecnico()
                tecnicoNuevo.nombre = nombre
                tecnicoNuevo.apellido =apellido
                tecnicoNuevo.cedula =cedula
                tecnicoNuevo.telefono =telefono
                tecnicoNuevo.huella = ""
                tecnicoNuevo.direccion = direccion
                rango = rangos(id=3)
                estado = estadoTecnico(id=1)
                tecnicoNuevo.estado = estado
                tecnicoNuevo.tipo_rango = rango
                tecnicoNuevo.save()
                return render(request,"modulos/admin/crearEmpleado.html",{
                    "mensaje1": "Operacion exitosa!"
                    })
        else:
                return render(request,"modulos/admin/crearEmpleado.html",{
                    "mensaje": "todos los campos son obligatorios"
                    })
    else:
        return render(request,"modulos/admin/crearEmpleado.html")
    
def crearSubAdmin(request):
    id_cli_sessi = request.session.get('identificador')
    try:
        compro_cli = admin.objects.get(id=id_cli_sessi)
        if compro_cli:
            pass
    except:
        return cerrarSesion_cantv(request)
    if request.method == "POST":
        nombre = request.POST.get("subAdmin")
        apellido = request.POST.get("apellidoSubAdmin")
        cedula = request.POST.get("cedulaSubAdmin")
        direccion = request.POST.get("direccion")
        telefono = request.POST.get("telefono")
        if nombre and apellido and cedula:
                subAdminNuevo = subAdmin()
                subAdminNuevo.nombre = nombre
                subAdminNuevo.apellido =apellido
                subAdminNuevo.cedula =cedula
                subAdminNuevo.direccion = direccion
                subAdminNuevo.telefono = telefono
                rango = rangos(id=2)
                subAdminNuevo.tipo_rango = rango
                subAdminNuevo.save()
                return render(request,"modulos/admin/crearSubAdmin.html",{
                    "mensaje1": "Operacion exitosa!"
                    })
        else:
                return render(request,"modulos/admin/crearSubAdmin.html",{
                    "mensaje": "todos los campos son obligatorios"
                    })
    else:
         return render(request,"modulos/admin/crearSubAdmin.html")
def eliminar_tecnico(request, id):
    eliminarTecnico = tecnico.objects.get(id=id)
    eliminarTecnico.eliminado=True
    eliminarTecnico.save()
    return redirect('admin')
def eliminar_subAdmin(request, id):
    eliminarSubAdmin = subAdmin.objects.get(id=id)
    eliminarSubAdmin.eliminado=True
    eliminarSubAdmin.save()
    return redirect('admin')
def eliminar_cliente(request, id):
    eliminarCliente = cliente.objects.get(id=id)
    eliminarCliente.eliminado=True
    eliminarCliente.save()
    return redirect('admin')
    #=======MODULO SUB ADMIN=====
def modulo_subAdmin(request):
    id_cli_sessi = request.session.get('identificador')
    algo_mal = None
    #**************************************************************
    sospecha = request.session.get('mensaje')
    if sospecha:
        # Eliminar el mensaje de la sesión para que no se muestre nuevamente después
        # de un nuevo renderizado.
        del request.session['mensaje']
    algo_mal = True if sospecha == 'cliente sospechoso' else False

    #***********************************************************************
    try:
        compro_cli = subAdmin.objects.get(id=id_cli_sessi)
        if compro_cli:
            pass
    except:
        return cerrarSesion_cantv(request)
    informacion = None
    mensaje = ""
    consultar_permiso1 = None
    consultar_permiso3 = None
    consultar_permiso4 = None
    consultar_permiso5 = None
    #=========INSTANCIAS======================
    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    instan_permiso_tipo1 = tiposPermisos.objects.get(id=1)
    instan_permiso_tipo3 = tiposPermisos.objects.get(id=3)
    instan_permiso_tipo4 = tiposPermisos.objects.get(id=4)
    instan_permiso_tipo5 = tiposPermisos.objects.get(id=5)
    #***********************************************************
    permisos_lista= {}
    # Intentamos obtener cada permiso y asignarlo a su clave correspondiente en el diccionario
    try:
        consultar_permiso1 = Permiso.objects.get(sub_admin=id_cli_sessi, tipo_permiso=instan_permiso_tipo1)
        if consultar_permiso1:
            permisos_lista["permiso1"] = consultar_permiso1
    except Permiso.DoesNotExist:
        permisos_lista["permiso1"] = None

    try:
        consultar_permiso3 = Permiso.objects.get(sub_admin=id_cli_sessi, tipo_permiso=instan_permiso_tipo3)
        if consultar_permiso3:
            permisos_lista["permiso3"] = consultar_permiso3
    except Permiso.DoesNotExist:
        permisos_lista["permiso3"] = None
    try:
        consultar_permiso4 = Permiso.objects.get(sub_admin=id_cli_sessi, tipo_permiso=instan_permiso_tipo4)
        if consultar_permiso4:
            permisos_lista["permiso4"] = consultar_permiso4
    except Permiso.DoesNotExist:
        permisos_lista["permiso4"] = None
    try:
        consultar_permiso5 = Permiso.objects.get(sub_admin=id_cli_sessi, tipo_permiso=instan_permiso_tipo5)
        if consultar_permiso5:
            permisos_lista["permiso5"] = consultar_permiso5
    except Permiso.DoesNotExist:
        permisos_lista["permiso5"] = None
    permisos_lista = {
                "permiso1":consultar_permiso1,
                "permiso2":consultar_permiso3,
                "permiso3":consultar_permiso4,
                "permiso4":consultar_permiso5,
            }
    #++++++++++++++++++++++++++++++++++++++++++++++++++++++
    clientesa = cliente.objects.filter(eliminado=False)
    empleados = tecnico.objects.filter(eliminado=False)
    disp = estadoTecnico.objects.get(id=1)
    ocup = estadoTecnico.objects.get(id=2)
    emple_report = reportesParaTecnico.objects.filter(status=4).count() == 5
    emple_report_status_3 = reportesParaTecnico.objects.filter(status=3).count() == 5
    #==========================================
    if request.method == 'POST':
        estadisticas_fallas_report = request.POST.get('estad_report_select')
        busqueda = request.POST.get('busqueda')
        resultados = []
        informacion = []
        modelos = [admin, tecnico, cliente, subAdmin]
         # Iterar en los modelos
        if busqueda:
            for modelo in modelos:
                modelo_results = modelo.objects.filter(
                    Q(nombre__icontains=busqueda) |    # Comparar por nombre
                    Q(cedula__icontains=busqueda) |     # Comparar por cédula
                    Q(apellido__icontains=busqueda)|    # apellido
                    Q(id__icontains=busqueda)          # Comparar por ID
                )
                resultados.extend(modelo_results)
            if not resultados:
                mensaje = "sin resultados"

            # Realizar las acciones necesarias con los resultados
            for resultado in resultados:
                data = {
                    "id": resultado.id,
                    "nombre":resultado.nombre,
                    "apellido":resultado.apellido,
                    "cedula":resultado.cedula,
                    "rango_id":resultado.tipo_rango.id,
                    "rango":resultado.tipo_rango.rango,
                }
                informacion.append(data)
        else:
            mensaje = "campo vacio"
    if not reportesParaTecnico.objects.exists() or emple_report or emple_report_status_3:
        tecnicos_asig = tecnico.objects.filter(estado=ocup if emple_report_status_3 else disp)
        tecnicos_asig.update(estado=disp)
        return render(request, 'modulos/subAdmin/subAdmin.html',{
        "clientes":clientesa,
        "empleados":empleados,
        "search_results":informacion,
        "sinResult":mensaje,
        "permisos":permisos_lista,
        'algo_mal': algo_mal
        })
    return render(request, 'modulos/subAdmin/subAdmin.html',{
        "clientes":clientesa,
        "empleados":empleados,
        "search_results":informacion,
        "sinResult":mensaje,
        "permisos":permisos_lista,
        'algo_mal': algo_mal
        })
def algo_anda_mal(request,id):
    request.session['mensaje'] = 'cliente sospechoso'
    return redirect("subAdmin")
def eliminar_cliente_sub(request, id):
    eliminarCliente = cliente.objects.get(id=id)
    eliminarCliente.eliminado=True
    eliminarCliente.save()
    return redirect('subAdmin')

def ver_reportes_gestion_sub(request):
    id_cli_sessi = request.session.get('identificador')
    try:
        compro_cli = subAdmin.objects.get(id=id_cli_sessi)
        if compro_cli:
            pass
    except:
        return cerrarSesion_cantv(request)
    instan_permiso_tipo6 = tiposPermisos.objects.get(id=6)
    reporteGestion = reportesParaTecnico.objects.filter(status=4,Datos_Cliente__isnull=False,DatosReporteCliente__isnull=False)
    reporteGestion1 = reportesParaTecnico.objects.filter(status=4,Datos_Cliente__isnull=True,DatosReporteCliente__isnull=True)
    # Combinar los resultados de ambas consultas
    reportes_combinados = list(reporteGestion) + list(reporteGestion1)
    if 'export_excel' in request.GET:
        # Crear un nuevo libro de Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Agregar encabezados de columna
        headers = ["ID", "Nombre Cliente", "Tipo de Falla", "Cédula Cliente", "Fecha Reporte", "Nombre Técnico", "Status", "Informacion de tecnico", "Fecha de gestion"]
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            sheet[f"{column_letter}1"] = header
            sheet[f"{column_letter}1"].font = Font(bold=True)

        # Agregar datos de los reportes gestionados
        for row_num, reporte in enumerate(reportes_combinados, 2):
            sheet[f"A{row_num}"] = reporte.id
            if reporte.Datos_Cliente:
                sheet[f"B{row_num}"] = reporte.Datos_Cliente.nombre
                sheet[f"D{row_num}"] = reporte.Datos_Cliente.cedula
                fecha_reporte = reporte.DatosReporteCliente.fecha if reporte.DatosReporteCliente.fecha else timezone.now()
                fecha_reporte_str = fecha_reporte.strftime("%Y-%m-%d %H:%M:%S")
                sheet[f"E{row_num}"] = fecha_reporte_str
            else:
                sheet[f"B{row_num}"] = reporte.DatosCliente_nombre
                sheet[f"D{row_num}"] = reporte.DatosCliente_cedula
                fecha_reporte = reporte.DatosCliente_FechaCreacion if reporte.DatosCliente_FechaCreacion else timezone.now()
                fecha_reporte_str = fecha_reporte.strftime("%Y-%m-%d %H:%M:%S")
                sheet[f"E{row_num}"] = fecha_reporte_str

            sheet[f"C{row_num}"] = reporte.DatosReporteCliente.categoria if reporte.DatosReporteCliente else reporte.DatosCliente_falla
            sheet[f"F{row_num}"] = reporte.DatosTecnico.nombre
            sheet[f"G{row_num}"] = reporte.status.estadoReporter
            sheet[f"H{row_num}"] = reporte.mas_informacion_tecnico if reporte.mas_informacion_tecnico else ""
            sheet[f"I{row_num}"] = reporte.fecha_gestion if reporte.fecha_gestion else timezone.now()
            fecha_gestion = reporte.fecha_gestion if reporte.fecha_gestion else timezone.now()
            fecha_gestion_str = fecha_gestion.strftime("%Y-%m-%d %H:%M:%S")  # Formato de cadena sin zona horaria
            sheet[f"I{row_num}"] = fecha_gestion_str

        # Ajustar el ancho de las columnas al contenido
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) if max_length > len(header) else len(header)
            column_letter = column[0].column_letter
            sheet.column_dimensions[column_letter].width = adjusted_width

        # Guardar el libro de Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=reportes_gestionados.xlsx'
        workbook.save(response)
        return response

    return render(request, 'modulos/subAdmin/paginasSubAdmin/reportesGestion_sub.html', {
        "reportesGestionados": reportes_combinados,
        "permiso4":instan_permiso_tipo6
    })
def ver_reportes_tecnico_sub(request, id):
    id_cli_sessi = request.session.get('identificador')
    try:
        compro_cli = subAdmin.objects.get(id=id_cli_sessi)
        if compro_cli:
            pass
    except:
        return cerrarSesion_cantv(request)
    tecnico_identidad = tecnico.objects.get(id=id)
    paraTecnicoo = reportesParaTecnico.objects.filter(DatosTecnico=tecnico_identidad, status=3)

    if request.method == "POST":
        reporte_reasignar_id = request.POST.get('reporte_id')
        reporte_reasignar = reportesParaTecnico.objects.get(id=reporte_reasignar_id)
        tecnico_disponible = tecnico.objects.filter(estado__estado='disponible').first()

        if tecnico_disponible:
            # Obtener el técnico anterior y actualizar su estado si no tiene más informes asignados
            informes_asignados_anterior = reportesParaTecnico.objects.filter(DatosTecnico=tecnico_identidad).count()
            if informes_asignados_anterior == 1:
                tecnico_identidad.estado = estadoTecnico.objects.get(id=1)
                tecnico_identidad.save()
            # Actualizar el informe para incluir el nuevo técnico
            reporte_reasignar.DatosTecnico = tecnico_disponible
            reporte_reasignar.save()

            # Actualizar el estado del técnico nuevo si tiene cinco informes asignados
            informes_asignados = reportesParaTecnico.objects.filter(DatosTecnico=tecnico_disponible).count()
            if informes_asignados == 5:
                tecnico_disponible.estado = estadoTecnico.objects.get(id=2)
                tecnico_disponible.save()
                return redirect("subAdmin")
            #===========================
    return render(request, 'modulos/subAdmin/ver_reportes_tecnico_sub.html', {
        "tecnico": tecnico_identidad, 
        "nuevo_reporte": paraTecnicoo
        })


def enviar_reporte_tecnico_sub(request, id):
    id_cli_sessi = request.session.get('identificador')
    try:
        compro_cli = subAdmin.objects.get(id=id_cli_sessi)
        if compro_cli:
            pass
    except:
        return cerrarSesion_cantv(request)
    reporte_cliente = reporteCliente.objects.filter(valor=True)
    tecnico_enviar = tecnico.objects.get(id=id)
    paraTecnicoo = reportesParaTecnico.objects.filter(DatosTecnico=tecnico_enviar,status=3)

    if request.method == 'POST':
        reporte_id = request.POST.get('reporte_id')
        tecnico_id = request.POST.get('tecnico_id')
        #=======instancias==========
        reporte_seleccionado = reporteCliente.objects.get(id=reporte_id)
        tecnico_seleccionado = tecnico.objects.get(id=tecnico_id)
        estado_reporte = estadoReporte.objects.get(estadoReporter="asignado")
        admin_sesion = request.session.get('cedula')
        subAdmin_id = get_object_or_404(subAdmin, cedula=admin_sesion)
        subAdmin_id = subAdmin_id.id
        subAdmin_instancia = subAdmin.objects.get(id=subAdmin_id)
        #============================
        informes_asignados = reportesParaTecnico.objects.filter(DatosTecnico=tecnico_seleccionado).count()
        if tecnico_seleccionado.estado.estado == "disponible" or informes_asignados < 5:
            el_tipo = tipoReporte.objects.get(id=1)
            nuevo_reporte = reportesParaTecnico(
                DatosCliente=reporte_seleccionado,
                DatosReporteCliente=reporte_seleccionado,
                status=estado_reporte,
                DatosTecnico=tecnico_seleccionado,
                DatosSubAdmin=subAdmin_instancia,
                tipo_reporte=el_tipo
            )
        informes_asignados += 1
        reporte_seleccionado.valor = False
        reporte_seleccionado.save()
        nuevo_reporte.save()
        if informes_asignados == 5:
            tecnico_seleccionado.estado.estado = "ocupado"
            tecnico_seleccionado.save()
            return redirect('subAdmin')
        else:
            return render(request, 'modulos/subAdmin/enviar_reporte_tecnico_sub.html',{"reporte":reporte_cliente,
             "tecnico":tecnico_enviar,
            "nuevo_reportee": paraTecnicoo
            })
    else:
        return render(request, 'modulos/subAdmin/enviar_reporte_tecnico_sub.html',{
            "reporte":reporte_cliente,
            "tecnico": tecnico_enviar, 
            "nuevo_reportee": paraTecnicoo
            })
            
def eliminar_tecnico_sub(request, id):
    eliminarTecnico = tecnico.objects.get(id=id)
    eliminarTecnico.eliminado=True
    eliminarTecnico.save()
    return redirect("subAdmin") 
def editarTecnico_sub(request, id):
    id_cli_sessi = request.session.get('identificador')
    try:
        compro_cli = subAdmin.objects.get(id=id_cli_sessi)
        if compro_cli:
            pass
    except:
        return cerrarSesion_cantv(request)
    tecnico_obj = tecnico.objects.get(id=id)
    empleado = tecnico.objects.filter(id=id)
    mensaje = ""

    if request.method == "POST":
        nombre = request.POST.get("nombre")
        apellido = request.POST.get("apellido")
        direccion = request.POST.get("direccion")
        telefono = request.POST.get("telefono")

        if nombre:
            tecnico_obj.nombre = nombre

        if apellido:
            tecnico_obj.apellido = apellido

        if direccion:
            tecnico_obj.direccion = direccion
        if telefono:
            tecnico_obj.telefono = telefono

        tecnico_obj.save()
        return redirect('subAdmin')

    return render(request, 'modulos/subAdmin/editarTecnico_sub.html', {
        "tecnico": tecnico_obj, 
        "empleados": empleado, 
        "estadoBusqueda": mensaje
        })

def reportesClientesSubAdmin(request):
    id_cli_sessi = request.session.get('identificador')
    try:
        compro_cli = subAdmin.objects.get(id=id_cli_sessi)
        if compro_cli:
            pass
    except:
        return cerrarSesion_cantv(request)
    historial = reporteCliente.objects.all()
    return render(request, 'modulos/subAdmin/paginasSubAdmin/reportes.html',{
        "historialCliente":historial
        })
def crearEmpleado_modulo_subAdmin(request):
    id_cli_sessi = request.session.get('identificador')
    try:
        compro_cli = subAdmin.objects.get(id=id_cli_sessi)
        if compro_cli:
            pass
    except:
        return cerrarSesion_cantv(request)
    if request.method == "POST":
        nombre = request.POST.get("empleado")
        apellido = request.POST.get("apellidoEmpleado")
        cedula = request.POST.get("cedulaEmpleado")
        direccion = request.POST.get("direccion")
        telefono = request.POST.get("telefono")
        if nombre and apellido and direccion and cedula:
                tecnicoNuevo = tecnico()
                tecnicoNuevo.nombre = nombre
                tecnicoNuevo.apellido =apellido
                tecnicoNuevo.cedula =cedula
                tecnicoNuevo.direccion =direccion
                tecnicoNuevo.telefono =telefono
                tecnicoNuevo.huella = ""
                rango = rangos(id=3)
                estado = estadoTecnico(id=1)
                tecnicoNuevo.estado = estado
                tecnicoNuevo.tipo_rango = rango
                tecnicoNuevo.save()
                return render(request,"modulos/subAdmin/crearEmpleado.html",{
                    "mensaje1": "Operacion exitosa!"
                    })
        else:
                return render(request,"modulos/subAdmin/crearEmpleado.html",{
                    "mensaje": "todos los campos son obligatorios"
                    })
    else:
         return render(request,"modulos/subAdmin/crearEmpleado.html")
    #======MODULO TECNICO========
def modulo_tecnico(request):
    id_cli_sessi = request.session.get('identificador')
    try:
        compro_cli = tecnico.objects.get(id=id_cli_sessi)
        if compro_cli:
            pass
    except:
        return cerrarSesion_cantv(request)
    tecnico_user= request.session.get('usuario')
    tecnico_name = tecnico(nombre=tecnico_user)
    return render(request, 'modulos/tecnico/empleado.html',{"empleado":tecnico_name})
def reportes_asignados(request):
    id_cli_sessi = request.session.get('identificador')
    try:
        compro_cli = tecnico.objects.get(id=id_cli_sessi)
        if compro_cli:
            pass
    except:
        return cerrarSesion_cantv(request)
    tecnico_id = request.session.get('identificador')
    mis_reportes = reportesParaTecnico.objects.filter(DatosTecnico=tecnico_id, status=3)
    if request.method == "POST":
        id_reporte = request.POST.get('id_reporte_asignado')
        audio_thread = threading.Thread(target=reproducir_audio_gestion_tecnico)
        audio_thread.start()
        return render(request, 'modulos/tecnico/formulario_gestion.html', {
            "id_reporte": id_reporte,
            "reportes": mis_reportes
            })
    else:
        return render(request, 'modulos/tecnico/reportes_asignados.html', {
            "reportes": mis_reportes
            })

def formulario_gestion(request):
    id_cli_sessi = request.session.get('identificador')
    try:
        compro_cli = tecnico.objects.get(id=id_cli_sessi)
        if compro_cli:
            pass
    except:
        return cerrarSesion_cantv(request)
    if request.method == "POST":
        id_reporte = request.POST.get('reporte_id')
        informacion = request.POST.get('informacion_adicional')
        fecha_gestiones = datetime.datetime.now()
        #====instancias=======
        instancia_estado = estadoReporte.objects.get(id=4)
        emple_dis = estadoTecnico.objects.get(id=1)
        #=========================
        gestion_datos = reportesParaTecnico.objects.get(id=id_reporte)
        gestion_datos.status= instancia_estado
        gestion_datos.mas_informacion_tecnico=informacion
        gestion_datos.fecha_gestion = fecha_gestiones
        gestion_datos.save()
#=======INSTANCIA PARA ELIMINAR QR Y IMAGEN TESTEP CLIENTE=========
        probar_cli_data_media = reportesParaTecnico.objects.get(status=instancia_estado,id=id_reporte)
        probar_cli_data_media.DatosReporteCliente.capture.delete()
        probar_cli_data_media.DatosReporteCliente.codigo_qr.delete()
        probar_cli_data_media.save()
#==============================================================
        tecnico_asig = gestion_datos.DatosTecnico
        if reportesParaTecnico.objects.filter(DatosTecnico=tecnico_asig,status=instancia_estado).count() == 5:
            tecnico_asig.estado = emple_dis
            tecnico_asig.save()

        #============
        actual = reportesParaTecnico.objects.get(id=id_reporte)
        actual.DatosReporteCliente.status = instancia_estado
        actual.DatosReporteCliente.save()
        actual.save()

        #============
        #==============================
        return redirect("reportesAsignados")
    else:
        return redirect("informeGestion")

    #=========MODULO CLIENTE========
def modulo_cliente(request):
    id_cli_sessi = request.session.get('identificador')
    try:
        compro_cli = cliente.objects.get(id=id_cli_sessi)
        if compro_cli:
            pass
    except:
        return cerrarSesion_user(request)
    cliente_session = request.session["usuario"]
    cliente_session_cedula = request.session["cedula"]
    datos_cliente = cliente.objects.filter(cedula=cliente_session_cedula)
    return render(request, 'modulos/cliente/cliente.html',{
        "datos_cliente": datos_cliente,
        "cliente":cliente_session,
        "datos_cliente":datos_cliente
        })

def cliente_reporte(request):
    id_cli_sessi = request.session.get('identificador')
    algo_mal = None
    #**************************************************************
    sospecha = request.session.get('mensaje')
    if sospecha:
        # Eliminar el mensaje de la sesión para que no se muestre nuevamente después
        # de un nuevo renderizado.
        del request.session['mensaje']
    algo_mal = True if sospecha == 'cliente sospechoso' else False
    consultar_permiso1 = None
    consultar_permiso3 = None
    consultar_permiso4 = None
    consultar_permiso5 = None
    #=========INSTANCIAS======================
    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    instan_permiso_tipo1 = tiposPermisos.objects.get(id=1)
    instan_permiso_tipo3 = tiposPermisos.objects.get(id=3)
    instan_permiso_tipo4 = tiposPermisos.objects.get(id=4)
    instan_permiso_tipo5 = tiposPermisos.objects.get(id=5)
    #***********************************************************
    permisos_lista= {}
    # Intentamos obtener cada permiso y asignarlo a su clave correspondiente en el diccionario
    try:
        consultar_permiso1 = Permiso.objects.get(cliente=id_cli_sessi, tipo_permiso=instan_permiso_tipo1)
        if consultar_permiso1:
            permisos_lista["permiso1"] = consultar_permiso1
    except Permiso.DoesNotExist:
        permisos_lista["permiso1"] = None

    try:
        consultar_permiso3 = Permiso.objects.get(cliente=id_cli_sessi, tipo_permiso=instan_permiso_tipo3)
        if consultar_permiso3:
            permisos_lista["permiso3"] = consultar_permiso3
    except Permiso.DoesNotExist:
        permisos_lista["permiso3"] = None
    try:
        consultar_permiso4 = Permiso.objects.get(cliente=id_cli_sessi, tipo_permiso=instan_permiso_tipo4)
        if consultar_permiso4:
            permisos_lista["permiso4"] = consultar_permiso4
    except Permiso.DoesNotExist:
        permisos_lista["permiso4"] = None
    try:
        consultar_permiso5 = Permiso.objects.get(cliente=id_cli_sessi, tipo_permiso=instan_permiso_tipo5)
        if consultar_permiso5:
            permisos_lista["permiso5"] = consultar_permiso5
    except Permiso.DoesNotExist:
        permisos_lista["permiso5"] = None
    permisos_lista = {
                "permiso1":consultar_permiso1,
                "permiso2":consultar_permiso3,
                "permiso3":consultar_permiso4,
                "permiso4":consultar_permiso5,
            }
    #***********************************************************************
    try:
        compro_cli = cliente.objects.get(id=id_cli_sessi)
        if compro_cli:
            pass
    except:
        return cerrarSesion_user(request)
    archivo_excel = 'C:/Users/antho/3D Objects/sistema_canTV/cantv_sistema/hojas_excel/fallas_de_cantv.xlsx'
    libro_excel = openpyxl.load_workbook(archivo_excel)
    hoja_excel = libro_excel['fallas de cantv']
    datos = []
    datos1 = []

    for fila in hoja_excel.iter_rows(min_row=2, values_only=True):
        falla = fila[0]
        if falla is not None:
            datos.append({'falla': falla})

    for filas in hoja_excel.iter_rows(min_row=2, values_only=True):
        informacion1 = filas[1]
        informacion2 = filas[2]
        if informacion1 is not None:
            datos1.append({'informacion1': informacion1, 'informacion2': informacion2})

    if request.method == 'POST':
        dato1 = request.POST.get('categoria')
        dato2 = request.POST.get('falla1')
        dato3 = request.POST.get('falla2')
        dato4 = request.FILES.get('prueba_velocidad')
        dato5 = request.POST.get('luces_pon')
        dato6 = request.POST.get('descripcion')

        if dato1 == "Seleccione falla":
            return render(request, 'modulos/cliente/generarReporte.html', {
                'datos': datos,
                  'datos1': datos1,
                    'mensaje': 'Debe seleccionar una opción válida'
                })
        elif dato1 == "Sin conexión a internet" and dato2 == "Indique la falla por desconexion de internet":
            return render(request, 'modulos/cliente/generarReporte.html', {
                'datos': datos, 
                'datos1': datos1, 
                'mensaje': 'Debe seleccionar la información detallada de la falla'
                })
        elif dato1 == "Hida y venida de internet" and dato3 == "Indique la falla por hida y venida de internet":
            return render(request, 'modulos/cliente/generarReporte.html', {
                'datos': datos, 
                'datos1': datos1, 
                'mensaje': 'Debe seleccionar la información detallada de la falla'
                })
        elif dato1 == "lento" and dato4 is None:
            return render(request, 'modulos/cliente/generarReporte.html', {
                'datos': datos, 
                'datos1': datos1, 'mensaje': 'Debe cargar una prueba de velocidad'
                })
        elif dato5 == "Seleccione la luz que vizualiza":
            return render(request, 'modulos/cliente/generarReporte.html', {
                'datos': datos, 
                'datos1': datos1, 
                'mensaje': 'Debe elegir que color de luz tiene su modem'
                })
        elif not dato6:
            return render(request, 'modulos/cliente/generarReporte.html', {
                'datos': datos, 
                'datos1': datos1, 
                'mensaje': 'Debe ingresar una descripcion'
                })
        else:
            sesion = request.session.get('cedula')
            cedula = cliente.objects.get(cedula=sesion)
            estado_reporte = estadoReporte.objects.get(estadoReporter="en proceso")
            estado_reporte_valor3 = estadoReporte.objects.get(id=3)
            estado_reporte_valor1 = estadoReporte.objects.get(id=1)
            tecnicos_disponibles = tecnico.objects.filter(estado=1,eliminado=False)

            if dato1 == "Sin conexión a internet":
                falla_reporte = dato2
                if tecnicos_disponibles.exists():
                    tecnico_asignado = tecnicos_disponibles.first()
                    el_tipo = tipoReporte.objects.get(id=1)
                    cliente_session_cedula = request.session["cedula"]
                    instan_cli = cliente.objects.get(cedula=cliente_session_cedula)

                    if reportesParaTecnico.objects.filter(DatosTecnico=tecnico_asignado).count() < 5:
                        reporteClientes = reporteCliente(
                            categoria=dato1,
                            falla=falla_reporte,
                            DatosCliente=cedula,
                            status=estado_reporte,
                            valor=False,
                            status_valor=estado_reporte_valor3,
                            capture=None,
                            luces=dato5,
                            description=dato6
                        )
                        reporteClientes.save()
                        
                        nuevo_reporte_tecnico = reportesParaTecnico(
                            Datos_Cliente=instan_cli,
                            DatosReporteCliente=reporteClientes,
                            DatosTecnico=tecnico_asignado,
                            status=estado_reporte_valor3,
                            fecha_gestion=None,
                            tipo_reporte=el_tipo
                        )
                        nuevo_reporte_tecnico.save()

                        if reportesParaTecnico.objects.filter(DatosTecnico=tecnico_asignado).count() == 5:
                            estado_ocupado = estadoTecnico.objects.get(id=2)
                            tecnico_asignado.estado = estado_ocupado
                            tecnico_asignado.save()
                    else:
                        estado_reporte = estadoReporte.objects.get(id=1)
                        reporteClientes = reporteCliente(
                            categoria=dato1,
                            falla=falla_reporte,
                            DatosCliente=cedula,
                            status=estado_reporte,
                            valor=False,
                            status_valor=estado_reporte_valor1,
                            capture=None,
                            luces=dato5,
                            description=dato6
                        )
                        reporteClientes.save()
                else:
                    estado_reporte = estadoReporte.objects.get(id=1)
                    reporteClientes = reporteCliente(
                            categoria=dato1,
                            falla=falla_reporte,
                            DatosCliente=cedula,
                            status=estado_reporte,
                            valor=False,
                            status_valor=estado_reporte_valor3,
                            capture=None,
                            luces=dato5,
                            description=dato6
                    )
                    reporteClientes.save()

            elif dato1 == "Servicio Intermitente":
                falla_reporte = dato3
                tecnicos_disponibles = tecnico.objects.filter(estado=1,eliminado=False)
                
                if tecnicos_disponibles.exists():
                    tecnico_asignado = tecnicos_disponibles.first()
                    el_tipo = tipoReporte.objects.get(id=1)
                    cliente_session_cedula = request.session["cedula"]
                    instan_cli = cliente.objects.get(cedula=cliente_session_cedula)

                    if reportesParaTecnico.objects.filter(DatosTecnico=tecnico_asignado).count() < 5:
                        reporteClientes = reporteCliente(
                            categoria=dato1,
                            falla=falla_reporte,
                            DatosCliente=cedula,
                            status=estado_reporte,
                            valor=False,
                            status_valor=estado_reporte_valor3,
                            capture=None,
                            luces=dato5,
                            description=dato6
                        )
                        reporteClientes.save()

                        nuevo_reporte_tecnico = reportesParaTecnico(
                            Datos_Cliente=instan_cli,
                            DatosReporteCliente=reporteClientes,
                            DatosTecnico=tecnico_asignado,
                            status=estado_reporte_valor3,
                            fecha_gestion=None,
                            tipo_reporte=el_tipo
                        )
                        nuevo_reporte_tecnico.save()

                        if reportesParaTecnico.objects.filter(DatosTecnico=tecnico_asignado).count() == 5:
                            estado_ocupado = estadoTecnico.objects.get(id=2)
                            tecnico_asignado.estado = estado_ocupado
                            tecnico_asignado.save()
                    else:
                        estado_reporte = estadoReporte.objects.get(id=1)
                        reporteClientes = reporteCliente(
                            categoria=dato1,
                            falla=falla_reporte,
                            DatosCliente=cedula,
                            status=estado_reporte,
                            valor=False,
                            status_valor=estado_reporte_valor1,
                            capture=None,
                            luces=dato5,
                            description=dato6
                        )
                        reporteClientes.save()
                else:
                    estado_reporte = estadoReporte.objects.get(id=1)
                    reporteClientes = reporteCliente(
                            categoria=dato1,
                            falla=falla_reporte,
                            DatosCliente=cedula,
                            status=estado_reporte,
                            valor=False,
                            status_valor=estado_reporte_valor1,
                            capture=None,
                            luces=dato5,
                            description=dato6
                    )
                    reporteClientes.save()

            else:
                falla_reporte = "internet lento"
                # ===========VALIDACION DE LA IMAGEN DE TESTEOD E VELOCIDAD DE INTERNET==========
                valid_formats = ['jpg', 'jpeg', 'png']

                # Verificar si la imagen cumple con los requisitos de formato y resolución
                try:
                    img_referencia = cv2.imread('media/imagen_validacion_testeo.png', 0)
                    img_nueva = cv2.imdecode(np.frombuffer(dato4.read(), np.uint8), cv2.IMREAD_GRAYSCALE)
                    img_nueva = cv2.resize(img_nueva, (img_referencia.shape[1], img_referencia.shape[0]))
                    img_format = dato4.name.split('.')[-1].lower()

                    if img_format not in valid_formats:
                        return render(request, 'modulos/cliente/generarReporte.html', {
                            'mensaje': 'El formato de imagen no es válido',
                            'datos': datos, 
                            'datos1': datos1
                            })
                    
                    ssim_score = ssim(img_referencia, img_nueva, full=True)[0]
                    umbral_similitud = 0.6

                    if ssim_score < umbral_similitud:
                        return render(request, 'modulos/cliente/generarReporte.html', {
                            'mensaje': 'La imagen no forma parte del testeo de cantv',
                            'datos': datos, 
                            'datos1': datos1
                            })
                    else:
                        tecnicos_disponibles = tecnico.objects.filter(estado=1,eliminado=False)
                        
                        if tecnicos_disponibles.exists():
                            tecnico_asignado = tecnicos_disponibles.first()
                            el_tipo = tipoReporte.objects.get(id=1)
                            cliente_session_cedula = request.session["cedula"]
                            instan_cli = cliente.objects.get(cedula=cliente_session_cedula)

                            if reportesParaTecnico.objects.filter(DatosTecnico=tecnico_asignado).count() < 5:
                                reporteClientes = reporteCliente(
                                    categoria=dato1,
                                    falla=falla_reporte,
                                    DatosCliente=cedula,
                                    status=estado_reporte,
                                    valor=False,
                                    status_valor=estado_reporte_valor3,
                                    capture=dato4,
                                    luces=dato5,
                                    description=dato6
                                )
                                reporteClientes.save()

                                nuevo_reporte_tecnico = reportesParaTecnico(
                                    Datos_Cliente=instan_cli,
                                    DatosReporteCliente=reporteClientes,
                                    DatosTecnico=tecnico_asignado,
                                    status=estado_reporte_valor3,
                                    fecha_gestion=None,
                                    tipo_reporte=el_tipo
                                )
                                nuevo_reporte_tecnico.save()

                                if reportesParaTecnico.objects.filter(DatosTecnico=tecnico_asignado).count() == 5:
                                    estado_ocupado = estadoTecnico.objects.get(id=2)
                                    tecnico_asignado.estado = estado_ocupado
                                    tecnico_asignado.save()
                            else:
                                estado_reporte = estadoReporte.objects.get(id=1)
                                reporteClientes = reporteCliente(
                                    categoria=dato1,
                                    falla=falla_reporte,
                                    DatosCliente=cedula,
                                    status=estado_reporte,
                                    valor=False,
                                    status_valor=estado_reporte_valor1,
                                    capture=dato4,
                                    luces=dato5,
                                    description=dato6
                                )
                                reporteClientes.save()
                        else:
                            estado_reporte = estadoReporte.objects.get(id=1)
                            reporteClientes = reporteCliente(
                                categoria=dato1,
                                falla=falla_reporte,
                                DatosCliente=cedula,
                                status=estado_reporte,
                                valor=False,
                                status_valor=estado_reporte_valor1,
                                capture=dato4,
                                luces=dato5,
                                description=dato6
                            )
                            reporteClientes.save()
                except Exception as e:
                    traceback.print_exc()
                    error_message = str(e)
                    return render(request, 'modulos/cliente/generarReporte.html', {
                        'mensaje': f'Error al procesar la imagen: {error_message}'
                        })

                
            # Generar el código QR con la información del reporte
            qr = qrcode.QRCode(version=1, box_size=10, border=5)
            qr.add_data(f"Cliente: {cedula}, Categoría: {dato1}, Falla: {falla_reporte}")
            qr.make(fit=True)
            qr_image = qr.make_image(fill='black', back_color='white')

            # Guardar el código QR en el modelo reporteCliente
            image_io = BytesIO()
            qr_image.save(image_io, format='PNG')
            qr_file = ContentFile(image_io.getvalue())
            reporteClientes.codigo_qr.save(f"{reporteClientes.id}.png", qr_file, save=True)

            # Redirigir al usuario a una página de éxito o a cualquier otra página deseada
            return render(request,'modulos/cliente/generarReporte.html',{
                'datos': datos, 
                'datos1': datos1, 
                'mensaje1': 'reporte enviado con exito',
                "permisos":permisos_lista,
                'algo_mal': algo_mal
                })

    return render(request, 'modulos/cliente/generarReporte.html', {
        'datos': datos, 
        'datos1': datos1,
        "permisos":permisos_lista,
        'algo_mal': algo_mal
        })

def historial_reporte(request):
    id_cli_sessi = request.session.get('identificador')
    try:
        compro_cli = cliente.objects.get(id=id_cli_sessi)
        if compro_cli:
            pass
    except:
        return cerrarSesion_user(request)

    consultar_permiso1 = None
    consultar_permiso3 = None
    consultar_permiso4 = None
    consultar_permiso5 = None
    #=========INSTANCIAS======================
    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    instan_permiso_tipo1 = tiposPermisos.objects.get(id=1)
    instan_permiso_tipo3 = tiposPermisos.objects.get(id=3)
    instan_permiso_tipo4 = tiposPermisos.objects.get(id=4)
    instan_permiso_tipo5 = tiposPermisos.objects.get(id=5)
    #***********************************************************
    permisos_lista= {}
    # Intentamos obtener cada permiso y asignarlo a su clave correspondiente en el diccionario
    try:
        consultar_permiso1 = Permiso.objects.get(cliente=id_cli_sessi, tipo_permiso=instan_permiso_tipo1)
        if consultar_permiso1:
            permisos_lista["permiso1"] = consultar_permiso1
    except Permiso.DoesNotExist:
        permisos_lista["permiso1"] = None

    try:
        consultar_permiso3 = Permiso.objects.get(cliente=id_cli_sessi, tipo_permiso=instan_permiso_tipo3)
        if consultar_permiso3:
            permisos_lista["permiso3"] = consultar_permiso3
    except Permiso.DoesNotExist:
        permisos_lista["permiso3"] = None
    try:
        consultar_permiso4 = Permiso.objects.get(cliente=id_cli_sessi, tipo_permiso=instan_permiso_tipo4)
        if consultar_permiso4:
            permisos_lista["permiso4"] = consultar_permiso4
    except Permiso.DoesNotExist:
        permisos_lista["permiso4"] = None
    try:
        consultar_permiso5 = Permiso.objects.get(cliente=id_cli_sessi, tipo_permiso=instan_permiso_tipo5)
        if consultar_permiso5:
            permisos_lista["permiso5"] = consultar_permiso5
    except Permiso.DoesNotExist:
        permisos_lista["permiso5"] = None
    permisos_lista = {
                "permiso1":consultar_permiso1,
                "permiso2":consultar_permiso3,
                "permiso3":consultar_permiso4,
                "permiso4":consultar_permiso5,
            }
    # Obtener el historial de reportes del cliente
    cedula = request.session.get('cedula')
    cedula_instancia = cliente.objects.get(cedula=cedula)
    historial = reporteCliente.objects.filter(eliminado=False,DatosCliente=cedula_instancia)
    return render(request, 'modulos/cliente/historialReportes.html',{
        'historial': historial,
        'cedula': cedula,
        "permisos":permisos_lista
        })

def generar_solicitudes(request):
    id_cli_sessi = request.session.get('identificador')
    try:
        compro_cli = cliente.objects.get(id=id_cli_sessi)
        if compro_cli:
            pass
    except:
        return cerrarSesion_user(request)

    consultar_permiso1 = None
    consultar_permiso3 = None
    consultar_permiso4 = None
    consultar_permiso5 = None
    #=========INSTANCIAS======================
    #++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    instan_permiso_tipo1 = tiposPermisos.objects.get(id=1)
    instan_permiso_tipo3 = tiposPermisos.objects.get(id=3)
    instan_permiso_tipo4 = tiposPermisos.objects.get(id=4)
    instan_permiso_tipo5 = tiposPermisos.objects.get(id=5)
    #***********************************************************
    permisos_lista= {}
    # Intentamos obtener cada permiso y asignarlo a su clave correspondiente en el diccionario
    try:
        consultar_permiso1 = Permiso.objects.get(cliente=id_cli_sessi, tipo_permiso=instan_permiso_tipo1)
        if consultar_permiso1:
            permisos_lista["permiso1"] = consultar_permiso1
    except Permiso.DoesNotExist:
        permisos_lista["permiso1"] = None

    try:
        consultar_permiso3 = Permiso.objects.get(cliente=id_cli_sessi, tipo_permiso=instan_permiso_tipo3)
        if consultar_permiso3:
            permisos_lista["permiso3"] = consultar_permiso3
    except Permiso.DoesNotExist:
        permisos_lista["permiso3"] = None
    try:
        consultar_permiso4 = Permiso.objects.get(cliente=id_cli_sessi, tipo_permiso=instan_permiso_tipo4)
        if consultar_permiso4:
            permisos_lista["permiso4"] = consultar_permiso4
    except Permiso.DoesNotExist:
        permisos_lista["permiso4"] = None
    try:
        consultar_permiso5 = Permiso.objects.get(cliente=id_cli_sessi, tipo_permiso=instan_permiso_tipo5)
        if consultar_permiso5:
            permisos_lista["permiso5"] = consultar_permiso5
    except Permiso.DoesNotExist:
        permisos_lista["permiso5"] = None
    permisos_lista = {
                "permiso1":consultar_permiso1,
                "permiso2":consultar_permiso3,
                "permiso3":consultar_permiso4,
                "permiso4":consultar_permiso5,
            }
    if request.method == 'POST':
        usuarioid = request.session.get("identificador")
        solicitud = request.POST.get('solicitud')
        instan_tipo = tipoReporte.objects.get(id=2)
        if not solicitud or solicitud == "Solicitud":
            return render(request, 'modulos/cliente/solicitudes.html', {
                'mensaje': 'Debe seleccionar una solicitud válida',
                "permisos":permisos_lista
                })

        if solicitud == "cambioServicio":
            plan = request.POST.get('plan')
            capturar_cli = cliente.objects.get(id=usuarioid)
            if not plan or plan == "Planes":
                return render(request, 'modulos/cliente/solicitudes.html', {
                    'mensaje': 'Debe seleccionar un tipo de plan válido',
                    "permisos":permisos_lista
                    })
            elif plan == "aba ultra 60mb":
                obtener_peticion = peticionesCliente.objects.create(
                    peticion = solicitud,
                    tema = plan,
                    datosCliente = capturar_cli
,                    tipo_de_reporte = instan_tipo
            )
            elif plan == "aba ultra 100mb":
                obtener_peticion = peticionesCliente.objects.create(
                    peticion = solicitud,
                    tema = plan,
                    datosCliente = capturar_cli
,                    tipo_de_reporte = instan_tipo
            )
            elif plan == "aba ultra 200mb":
                obtener_peticion = peticionesCliente.objects.create(
                    peticion = solicitud,
                    tema = plan,
                    datosCliente = capturar_cli
,                    tipo_de_reporte = instan_tipo
            )
            elif plan == "aba ultra 300mb":
                obtener_peticion = peticionesCliente.objects.create(
                    peticion = solicitud,
                    tema = plan,
                    datosCliente = capturar_cli
,                    tipo_de_reporte = instan_tipo
            )
            return render(request, 'modulos/cliente/solicitudes.html',{
                "mensaje1":"Peticion enviada con exito",
                "permisos":permisos_lista
                })

        if solicitud == "Mudanza":
            direccion_nueva = request.POST.get('nueva_dic')
            if not direccion_nueva:
                return render(request, 'modulos/cliente/solicitudes.html', {
                    'mensaje': 'Debe ingresar una dirección nueva',
                    "permisos":permisos_lista
                    })
            else:
                capturar_cli = cliente.objects.get(id=usuarioid)
                obtener_peticion = peticionesCliente(
                    peticion = solicitud,
                    tema = direccion_nueva,
                    datosCliente = capturar_cli
,                    tipo_de_reporte = instan_tipo
                )
                obtener_peticion.save()
                return render(request, 'modulos/cliente/solicitudes.html',{
                    "mensaje1":"Peticion enviada con exito",
                    "permisos":permisos_lista
                    })
        if solicitud == "retiroServicio":
            motivo_retiro = request.POST.get('motivoRetiro')
            if not motivo_retiro:
                return render(request, 'modulos/cliente/solicitudes.html', {
                    'mensaje': 'Debe ingresar un motivo de retiro',
                    "permisos":permisos_lista
                    })
            else:
                capturar_cli = cliente.objects.get(id=usuarioid)
                obtener_peticion = peticionesCliente(
                    peticion = solicitud,
                    tema = motivo_retiro,
                    datosCliente = capturar_cli,
                    tipo_de_reporte = instan_tipo
                )
                obtener_peticion.save()
                return render(request, 'modulos/cliente/solicitudes.html',{
                    "mensaje1":"Peticion enviada con exito",
                    "permisos":permisos_lista
                    })

    return render(request, 'modulos/cliente/solicitudes.html',{
        "permisos":permisos_lista
    })


def historial_reporte_pdf(request):
    # Obtener el historial de reportes del cliente
    historial = reporteCliente.objects.filter(eliminado=False)
    cedula = request.session.get('cedula')

    # Crear el objeto PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="historial_reporte.pdf"'

    # Crear el lienzo PDF en orientación horizontal (landscape)
    pdf = canvas.Canvas(response, pagesize=landscape(letter))
    pdf.setTitle('Historial de reportes')

    # Configurar el tamaño y margen del lienzo
    width, height = landscape(letter)
    margin = 0.5 * inch
    usable_width = width - 2 * margin
    usable_height = height - 2 * margin

    # Agregar el encabezado al PDF
    pdf.setFont('Helvetica-Bold', 16)
    pdf.drawString(margin, height - margin - 0.5*inch, 'Historial de reportes del cliente con cédula {}'.format(cedula))

    # Agregar los datos del historial al PDF
    pdf.setFont('Helvetica', 12)
    y = height - margin - 1.5*inch
    for historiales in historial:
        pdf.drawString(margin, y, 'ID: {}'.format(historiales.id))
        pdf.drawString(margin, y-0.2*inch, 'Categoría: {}'.format(historiales.categoria))
        pdf.drawString(margin, y-0.4*inch, 'Descripción: {}'.format(historiales.falla))
        pdf.drawString(margin, y-0.6*inch, 'Fecha: {}'.format(historiales.fecha.strftime('%d/%m/%Y')))
        y -= 1*inch

        # Verificar si se alcanzó el final de la página
        if y <= margin:
            # Agregar una nueva página al PDF
            pdf.showPage()
            # Reiniciar el valor de 'y' para empezar en la nueva página
            y = height - margin - 1.5*inch

    # Guardar y cerrar el PDF
    pdf.save()

    # Devolver la respuesta HTTP con el PDF
    return response
def eliminar_reporte(request, id):
    reporte = reporteCliente.objects.get(id=id)
    reporte.eliminado=True
    reporte.save()
    return redirect('historialReporteCliente')
def eliminar_todos_reportes(request, cedula):
    reporteCliente.objects.filter(DatosCliente__cedula=cedula).update(eliminado=True)
    return redirect('historialReporteCliente')




